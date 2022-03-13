from django.http import HttpResponseNotFound, HttpResponseBadRequest, HttpResponse, JsonResponse, HttpRequest
from django.shortcuts import render

from rest_framework import viewsets
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView

from WebApp.models import Nation, League, Participant, Team, Position, Player, Group, Match
from WebApp.serializers import NationSerializer, LeagueSerializer, ParticipantSerializer, TeamSerializer, \
    PositionSerializer, PlayerSerializer, GroupSerializer, MatchSerializer
from WebApp.services import get_players_db, get_players_by_position_string, write_players_sheet, strip_player_positions_string, get_players_by_position_algorithm, update_teams_leagues
from django_filters.rest_framework import DjangoFilterBackend

import openpyxl

import operator


class NationViewSet(viewsets.ModelViewSet):
    queryset = Nation.objects.all()
    serializer_class = NationSerializer

    def create(self, request, *args, **kwargs):
        data = dict(request.data)
        nation = Nation.create(data['name'], data['image_path'])
        nation = NationSerializer(nation, context={'request': request})
        return Response(nation.data)


class LeagueViewSet(viewsets.ModelViewSet):
    queryset = League.objects.all()
    serializer_class = LeagueSerializer

    def create(self, request, *args, **kwargs):
        data = dict(request.data)
        league = League.create(data['name'])
        league = LeagueSerializer(league, context={'request': request})
        return Response(league.data)


class ParticipantViewSet(viewsets.ModelViewSet):
    queryset = Participant.objects.all()
    serializer_class = ParticipantSerializer

    def create(self, request, *args, **kwargs):
        data = dict(request.data)
        try:
            team = Team.objects.get(id=int(data['team']))
        except:
            return HttpResponseNotFound("Selected team not found")
        if team.participant != None:
            return HttpResponseBadRequest("The chosen team has already been assigned to another participant.")
        participant = Participant.create(
            data['name'], int(data['budget']), team)
        participant = ParticipantSerializer(
            participant, context={'request': request})
        return Response(participant.data)

    def partial_update(self, request, *args, **kwargs):
        try:
            instance = Participant.objects.get(id=kwargs.get('pk'))
        except:
            return HttpResponseNotFound("Participant not found")
        data = dict(request.data)
        team = data.get('team')
        if team != None:
            old_team = Team.objects.filter(participant=instance.id)
            new_team = Team.objects.filter(id=int(team))
            if len(new_team)>0 and (new_team[0].participant != None and new_team[0].participant.id != instance.id):
                    return HttpResponseBadRequest("The chosen team has already been assigned to another participant.")
            elif len(new_team) > 0:
                new_team = new_team[0]
            else:
                return HttpResponseBadRequest("The chosen team does not exist.")

            if len(old_team) > 0:
                old_team = old_team[0]

                players_old_team = Player.objects.filter(team_participant=old_team.id)
                for player in players_old_team:
                    player.team_participant = new_team
                    player.save()
                
                if team != old_team.id:
                    old_team.participant = None
       
            new_team.participant = instance
            if old_team:
                old_team.save()
            new_team.save()
            del data['team']

        serializer = self.serializer_class(instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        try:
            instance = Participant.objects.get(id=kwargs.get('pk'))
        except:
            return HttpResponseNotFound("Participant not found")
        team = Team.objects.get(participant=instance.id)
        players = Player.objects.filter(team_participant=team.id)
        for player in players:
            player.team_participant = None
            player.value = None
            player.save()
        instance.delete()
        return Response(ParticipantSerializer(instance).data)


class TeamViewSet(viewsets.ModelViewSet):
    queryset = Team.objects.all().order_by('name')
    serializer_class = TeamSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['league']

    def create(self, request, *args, **kwargs):
        data = dict(request.data)
        try:
            league = League.objects.get(id=int(data['league']))
        except:
            return HttpResponseNotFound("League not found")
        team = Team.create(data['name'], league, data['image_path'])
        team = TeamSerializer(team, context={'request': request})
        return Response(team.data)


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer

    def create(self, request, *args, **kwargs):
        data = dict(request.data)
        position = Position.create(data['name'])
        position = PositionSerializer(position, context={'request': request})
        return Response(position.data)


class PlayerViewSet(viewsets.ModelViewSet):
    queryset = Player.objects.all().order_by('-overall', '-pace')
    serializer_class = PlayerSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['position', 'team_participant', 'team_origin']

    def cut_player_quantity(self, position):
        n_participants = len(Participant.objects.all())
        if position.name == 'Goalkeepers':
            return get_players_by_position_algorithm(position.id, 1.5*n_participants)
        elif position.name == 'Center Backs':
            return get_players_by_position_algorithm(position.id, 3*n_participants)
        elif position.name == 'Full Backs':
            return get_players_by_position_algorithm(position.id, 3*n_participants)
        elif position.name == 'Defensive Midfielders':
            return get_players_by_position_algorithm(position.id, 3*n_participants)
        elif position.name == 'Ofensive Midfielders':
            return get_players_by_position_algorithm(position.id, 1.5*n_participants)
        elif position.name == 'Wingers':
            return get_players_by_position_algorithm(position.id, 2.5*n_participants)
        elif position.name == 'Attackers':
            return get_players_by_position_algorithm(position.id, 2*n_participants)

    def create(self, request, *args, **kwargs):
        data = dict(request.data)
        try:
            position = Position.objects.get(id=int(data['position']))
        except:
            return HttpResponseNotFound("Position not found")
        try:
            team_origin = Team.objects.get(id=int(data['team_origin']))
        except:
            return HttpResponseNotFound("Team not found")
        try:
            nation = Nation.objects.get(id=int(data['nation']))
        except:
            return HttpResponseNotFound("Nation not found")
        player = Player.create(data['name'], int(
            data['overall']), position, team_origin, nation, data['image_path'], data['specific_position'])
        player = PlayerSerializer(player, context={'request': request})
        return Response(player.data)

    def list(self, request, *args, **kwargs):
        players = list(Player.objects.all().order_by('-overall'))
        if 'team_participant' in request._request.__dict__['environ']['QUERY_STRING']:
            team_participant = request._request.__dict__['environ']['QUERY_STRING'].split('=')[1]
            players = [player for player in players if str(player.team_participant_id)==team_participant]
            players.sort(key=lambda item: (item.overall, item.pace), reverse=True)
        if 'position' in request._request.__dict__['environ']['QUERY_STRING']:
            position = request._request.__dict__['environ']['QUERY_STRING'].split('=')[1]
            try:
                position_obj = Position.objects.get(id=position)
            except:
                return HttpResponseNotFound("Position not found")
            players = self.cut_player_quantity(position_obj)
        
        return Response(PlayerSerializer(players, many=True).data)


class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer

    def create(self, request, *args, **kwargs):
        data = dict(request.data)
        group = Group.create(data['group'])
        group = GroupSerializer(group, context={'request': request})
        return Response(group.data)


class MatchViewSet(viewsets.ModelViewSet):
    queryset = Match.objects.all()
    serializer_class = MatchSerializer

    def create(self, request, *args, **kwargs):
        data = dict(request.data)
        try:
            group = Group.objects.get(id=int(data['group']))
        except:
            return HttpResponseNotFound("Group not found")
        try:
            participant_1 = Participant.objects.get(id=int(data['participant_1']))
        except:
            return HttpResponseNotFound("Selected participant 1 not found")
        try:
            participant_2 = Participant.objects.get(id=int(data['participant_2']))
        except:
            return HttpResponseNotFound("Selected participant 2 not found")
        match = Match.create(group, participant_1, participant_2)
        match = MatchSerializer(match, context={'request': request})
        return Response(match.data)


class BuyPlayerView(APIView):
    def post(self, request, id):
        player = Player.objects.filter(id=id)
        if len(player) == 0:
            return HttpResponseNotFound("There are no players with the given id.")
        else:
            player = player[0]

        data = dict(request.data)

        team = None
        price = None
        if data.get('team_participant') is not None:
            team = Team.objects.filter(id=int(data['team_participant']))
            if len(team) == 0:
                return HttpResponseNotFound("There are no teams with the given id.")
            else:
                team = team[0]
            price = int(data['value'])

        result = player.buy_player(team, price)

        if result:
            return Response(PlayerSerializer(player).data)
        else:
            return HttpResponseBadRequest("The budget of {} is not enough to buy {}".format(data['team_participant'], player))

class GetGroupStageView(APIView):
    def post(self, request):
        data = dict(request.data)
        groups = Group.objects.all()
        if len(groups) > 0 and not data['replace']:
            return Response(GroupSerializer(groups[0]).data)
        elif len(groups) > 0:
            for group in groups:
                group.delete()
        
        participants = Participant.objects.all()
        group = Group.create('Group 1')
        for participant in participants:
            group.add_participant(participant)
        
        group.create_matches()

        group.export_matches()

        return Response(GroupSerializer(group).data)

class GetWildcardKnockoutRoundView(APIView):
    def post(self, request):
        data = dict(request.data)
        matches_wildcard = []

        old_stage = Group.objects.filter(group='Wildcard')
        if len(old_stage) > 0 and not data['replace']:
            return Response(GroupSerializer(old_stage[0]).data)
        elif len(old_stage) > 0:
            old_stage = old_stage[0]
            old_stage.delete()

        groups = list(Group.objects.filter(group='Group 1'))

        if len(groups) == 1:
            table_wildcard = groups[0].get_group_table()[2:6]
            wildcard = Group.create('Wildcard')

            for participant in table_wildcard:
                wildcard.add_participant(participant[0])

            matches_wildcard.append(Match.create(wildcard, table_wildcard[0][0], table_wildcard[3][0]))
            matches_wildcard.append(Match.create(wildcard, table_wildcard[1][0], table_wildcard[2][0]))

            wildcard.matches = matches_wildcard
            
            wildcard.save()
        else:
            return HttpResponseBadRequest("Number of groups does not allow the generation of a knockout round.")

        return Response(GroupSerializer(wildcard).data)


class GetSemiFinalsRoundView(APIView):
    def check_position_group_stage(self, participant, group_stage):
        for i, position in enumerate(group_stage):
            if position[0] == participant:
                return i
        return -1

    def get_group_winners(self, group, participants, group_stage):
        qualified = []
        try:
            group = Group.objects.get(group=group)
        except:
            return HttpResponseNotFound("Group not found")
        for participant in participants:
            if participant[1]['P'] == 3:
                qualified.append(participant[0])
            elif participant[1]['P'] == 1:
                opponent = group.find_match_group_participant_opponent(participant[0])
                position_participant = self.check_position_group_stage(participant[0], group_stage)
                position_opponent = self.check_position_group_stage(opponent, group_stage)
                if position_participant < position_opponent:
                    qualified.append(participant[0])

            if len(qualified) == 2:
                break

        return qualified

    def post(self, request):
        data = dict(request.data)
        matches_semis = []

        old_stage = Group.objects.filter(group='Semis')
        if len(old_stage) > 0 and not data['replace']:
            return Response(GroupSerializer(old_stage[0]).data)
        elif len(old_stage) > 0:
            old_stage = old_stage[0]
            old_stage.delete()

        try:
            group_stage = Group.objects.get(group='Group 1')
        except:
            return HttpResponseNotFound("Group stage not found")
        table_group_stage = group_stage.get_group_table()
 
        try:
            wildcard_stage = Group.objects.get(group='Wildcard')
        except:
            return HttpResponseNotFound("Wildcard stage not found")
        table_wildcard = wildcard_stage.get_group_table()
        semis = Group.create('Semis')

        wildcard_winners = self.get_group_winners('Wildcard', table_wildcard, table_group_stage)

        semis.add_participant(table_group_stage[0][0])
        semis.add_participant(table_group_stage[1][0])
        semis.add_participant(wildcard_winners[0])
        semis.add_participant(wildcard_winners[1])

        wc1_position = self.check_position_group_stage(wildcard_winners[0], table_group_stage)
        wc2_position = self.check_position_group_stage(wildcard_winners[1], table_group_stage)

        if wc1_position < wc2_position:
            matches_semis.append(Match.create(semis, table_group_stage[0][0], wildcard_winners[1]))
            matches_semis.append(Match.create(semis,  wildcard_winners[1], table_group_stage[0][0]))
            matches_semis.append(Match.create(semis, table_group_stage[1][0], wildcard_winners[0]))
            matches_semis.append(Match.create(semis, wildcard_winners[0], table_group_stage[1][0]))
        else:
            matches_semis.append(Match.create(semis, table_group_stage[0][0], wildcard_winners[0]))
            matches_semis.append(Match.create(semis, wildcard_winners[0], table_group_stage[0][0]))
            matches_semis.append(Match.create(semis, table_group_stage[1][0], wildcard_winners[1]))
            matches_semis.append(Match.create(semis, wildcard_winners[1], table_group_stage[1][0]))

        semis.matches = matches_semis

        semis.save()

        return Response(GroupSerializer(semis).data)


class GetFinalView(APIView):
    def check_position_group_stage(self, participant, group_stage):
        for i, position in enumerate(group_stage):
            if position[0] == participant:
                return i
        return -1
        
    def post(self, request):
        data = dict(request.data)
        
        old_stage = Group.objects.filter(group='Final')
        if len(old_stage) > 0 and not data['replace']:
            return Response(GroupSerializer(old_stage[0]).data)
        elif len(old_stage) > 0:
            old_stage = old_stage[0]
            old_stage.delete()

        semis = Group.objects.get(group='Semis')

        participants = semis.get_group_table()
        qualified = list()

        qualified.append(participants[0][0])
        qualified.append(participants[1][0])

        matches = []
            
        final = Group.create('Final')

        for participant in qualified:
            final.add_participant(participant)

        matches.append(Match.create(final, qualified[0], qualified[1]))

        final.matches = matches

        final.save()

        return Response(GroupSerializer(final).data)


class TransfersView(APIView):
    def get(self, request):
        file = open("transfers.txt", "w", encoding="utf-8")
        leagues = League.objects.all().order_by('name')
        team_participant = None
        for league in leagues:
            team_participant = None
            strBuilder = ""
            strBuilder += ("LEAGUE: {}\n".format(league.name.upper()))
            teams = Team.objects.filter(league=league.id).order_by('name')
            for team in teams:
                strBuilder += ("\tTEAM: {}\n".format(team.name.replace('ş', 's').upper()))
                players = Player.objects.filter(team_origin=team.id).order_by(
                    '-team_participant', '-overall')
                for player in players:
                    if player.team_participant != None and player.team_participant != team:
                        team_participant = player.team_participant
                        strBuilder += ("\t\tTO {}: {} - {}\n".format(
                            team_participant.name.upper(), player.overall, player.name.upper()))
            strBuilder += ('--------------------------------------------------------------------------\n')
            if team_participant is not None:
                file.write(strBuilder)

        file.close()

        return Response()

class ChampzPlayersView(APIView):
    def get(self, request):
        positions = list(Position.objects.all())
        wb = openpyxl.Workbook()
        for i,position in enumerate(positions):
            wb.create_sheet(index = i , title = position.name)
        n_participants = len(list(Participant.objects.all()))
        wb.active = 0
        get_players_by_position_string(wb.active, positions[0], 1.5*n_participants)
        wb.active = 1
        get_players_by_position_string(wb.active, positions[1], 3*n_participants)
        wb.active = 2
        get_players_by_position_string(wb.active, positions[2], 3*n_participants)
        wb.active = 3
        get_players_by_position_string(wb.active, positions[3], 3*n_participants)
        wb.active = 4
        get_players_by_position_string(wb.active, positions[4], 1.5*n_participants)
        wb.active = 5
        get_players_by_position_string(wb.active, positions[5], 2.5*n_participants)
        wb.active = 6
        get_players_by_position_string(wb.active, positions[6], 2*n_participants)

        wb.save("players.xlsx") 

        return Response()

class ParticipantsTeamsView(APIView):
    def get(self, request):
        participants = Participant.objects.all()
        wb = openpyxl.Workbook()
        for i,participant in enumerate(participants):
            participant_team = Team.objects.get(participant=participant)
            participant_players = Player.objects.filter(team_participant=participant_team).order_by('position', 'specific_position', '-overall')
            wb.create_sheet(index = i , title = "{}({})".format(participant.name, participant_team.name))
            wb.active = i
            write_players_sheet(wb.active, participant_players)

        wb.save("teams.xlsx") 

        return Response()

class EndChampzView(APIView):
    def post(self, request):
        file = open("champz.txt", "w", encoding="utf-8")
        participants = Participant.objects.all()
        file.write("TEAMS: \n\n")
        for participant in participants:
            strBuilder = ""
            strBuilder += 'TEAM {}'.format(participant.name.upper())
            team_participant = Team.objects.get(participant=participant)
            strBuilder += '({})\n'.format(team_participant.name.upper())
            players_participant = Player.objects.filter(team_participant=team_participant).order_by('-overall')
            for player in players_participant:
                strBuilder += '\t{} - {} - ${}\n'.format(player.name, player.overall, player.value)
            strBuilder += '\n'
            file.write(strBuilder)

        groups = list(Group.objects.all())
        group_stage = [g for g in groups if 'Group' in g.group]

        for index, group in enumerate(group_stage):
            table = group.get_group_table()

            file.write('GROUP {} TABLE\n\n'.format(index+1))
            file.write('TEAM\t\tP\tW\tD\tL\tGF\tGA\tGD\n')
            for el in table:
                participant = Participant.objects.get(name=el[0])
                if len(participant.name) > 8:
                    file.write('{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\n'.format(participant.name, el[1]['P'], el[1]['W'], el[1]['D'], el[1]['L'], el[1]['GF'], el[1]['GA'], el[1]['GD']))
                else:
                    file.write('{}\t\t{}\t{}\t{}\t{}\t{}\t{}\t{}\n'.format(participant.name, el[1]['P'], el[1]['W'], el[1]['D'], el[1]['L'], el[1]['GF'], el[1]['GA'], el[1]['GD']))
            file.write('\n')


        file.write('MATCHES: \n\n')

        for group in groups:
            strBuilder = ""
            strBuilder += '{}\n'.format(group.group.upper())
            matches = Match.objects.filter(group=group)
            for match in matches:
                participant_1 = match.participant_1.participant.name
                participant_2 = match.participant_2.participant.name
                goals_1 = match.goals_participant_1
                if goals_1 is None:
                    goals_1 = ''
                goals_2 = match.goals_participant_2
                if goals_2 is None:
                    goals_2 = ''
                strBuilder += '\t{} {} x {} {}\n'.format(participant_1, goals_1, goals_2, participant_2)
            strBuilder += '\n'
            file.write(strBuilder)

        file.close()

        return Response()

class UpdatePlayerDatabaseView(APIView):
    def post(self, request):
        get_players_db()
        return Response()

class UpdateTeamsLeaguesDatabaseView(APIView):
    def post(self, request):
        update_teams_leagues()
        return Response()
