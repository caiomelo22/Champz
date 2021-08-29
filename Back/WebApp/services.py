from WebApp.models import Player, Nation, Team, League, Position
from urllib.request import Request, urlopen as uReq
from bs4 import BeautifulSoup as soup
from selenium import webdriver
import time
from selenium.webdriver.chrome.options import Options
from openpyxl.styles import Color, PatternFill, Font, Border
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.styles import colors

def get_id_from_href(link):
    start = link.find("player/")
    end = link.find('/', start + 8)
    return link[start + 7:end]


# Verifica se alguma versão antiga daquele jogador já existe na lista e se existir, devolverá o index dele
def get_jogador_index(obj, position):
    players = list(Player.objects.filter(position=position))
    for player in players:
        if obj.name == player.name:
            return player.id
    return -1

def strip_player_positions_string():
    players = Player.objects.all()
    for player in players:
        player.specific_position = player.specific_position.split('|')[0].strip()
        player.save()

def get_futbin_data():
    positions = {'Goalkeepers': 'GK', 'Center Backs': 'CB', 'Full Backs': 'RB,LB', 'Defensive Midfielders': 'CDM,CM',
                 'Ofensive Midfielders': 'CAM', 'Wingers': 'LW,LF,LM,RF,RW,RM', 'Attackers': 'ST,CF'}
    # positions = {'Goalkeepers': 'GK'}

    base_url = "https://www.fifacm.com/players?position="

    option = Options()
    option.headless = True
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=option)
    # driver.maximize_window()
    # print(playerRows)

    # CRIANDO O LOOP PARA PASSAR POR TODAS AS POSICOES
    for name, positions in positions.items():
        query = Position.objects.filter(name=name)
        if len(query) == 0:
            position = Position.create(name)
        else:
            position = query[0]

        # PEGANDO OS JOGADORES DO FUTBIN

        for pagina in range(1,4):
            url = base_url + positions

            url = url + '&page=' + str(pagina)
            print(url)

            driver.get(url)
            time.sleep(5)

            igs_btns = driver.find_elements_by_class_name("igs-btn")
            scroll = 0
            cont = 0
            scrollStride = 450
            if positions == 'GK':
                scrollStride = 210

            for btn in igs_btns:
                btn.click()
                driver.execute_script("window.scrollTo({}, {})".format(scroll, scroll + scrollStride))
                scroll += scrollStride
                time.sleep(1)

                element = driver.find_elements_by_class_name("site-players-page")[0]
                htmlContent = element.get_attribute('outerHTML')

                page_soup = soup(htmlContent, "html.parser")

                table = page_soup.find("table")
                trList = table.findAll("tr")[1+cont:1+cont+2]

                # Basic Info
                player = Player(position=position)

                tds = trList[0].findAll("td")
                player.id = tds[0].find('div', {"class": "igs-btn"})['data-playerid']

                team = tds[1].findAll('img', {"class": "team-img"})[0]['data-original-title']
                nation = tds[1].findAll('img', {"class": "team-img"})[1]['data-original-title']
                team_img = tds[1].findAll('img', {"class": "team-img"})[0]['src']
                nation_img = tds[1].findAll('img', {"class": "team-img"})[1]['src']

                query = Team.objects.filter(name=team)
                if len(query) == 0:
                    team = Team.objects.create(name=team, image_link=team_img)
                else:
                    team = query[0]
                player.team_origin = team

                query = Nation.objects.filter(name=nation)
                if len(query) == 0:
                    nation = Nation.create(nation, nation_img)
                else:
                    nation = query[0]
                player.nation = nation

                player.specific_position = ''.join(tds[1].find('div', {"class": "player-position-cln"}).findAll(text=True)).split(',')[0].split('|').strip()
                stats = trList[1].find("div", {"class": "player-stats"}).findAll("div", {"class": "col-md-2"})
                player.image_link = tds[1].find('img', {"class": "player-img-info"})['src']
                player.name = ''.join(tds[1].find('div', {"class": "player-name"}).find("a").findAll(text=True))
                player.overall = ''.join(tds[2].find('div', {"class": "player-overall"}).findAll(text=True))
                player.pace = ''.join(stats[0].find("div", {"class": "main-stat-rating-title"}).findAll(text=True))
                player.shooting = ''.join(stats[1].find("div", {"class": "main-stat-rating-title"}).findAll(text=True))
                player.passing = ''.join(stats[2].find("div", {"class": "main-stat-rating-title"}).findAll(text=True))
                player.dribbling = ''.join(stats[3].find("div", {"class": "main-stat-rating-title"}).findAll(text=True))
                player.defending = ''.join(stats[4].find("div", {"class": "main-stat-rating-title"}).findAll(text=True))
                player.physical = ''.join(stats[5].find("div", {"class": "main-stat-rating-title"}).findAll(text=True))

                other_player = Player.objects.filter(id=player.id)
                if not other_player.exists():
                    player.save()

                cont += 2

            # PEGANDO DADOS DOS JOGADORES

            # cont = 0
            # lista = 1
            #
            # while cont < len(players_1):
            #     if lista == 1:
            #         if cont > len(players_1)-1:
            #             break
            #         jogador = players_1[cont]
            #     else:
            #         if cont > len(players_2)-1:
            #             break
            #         jogador = players_2[cont]
            #
            #     tds = jogador.findAll('td')
            #
            #     data = tds[0].findAll('a')
            #
            #     player = Player(position=position)
            #
            #     player_img = tds[0].find("div", {"class": "d-inline"})
            #     player_img = player_img.find('img')['data-original']
            #     player.image_link = player_img
            #
            #     player.name = ''.join(data[0].findAll(text=True))
            #
            #     player.id = get_id_from_href(data[0]['href'])
            #
            #     league = data[3]['data-original-title']
            #
            #     allowedLeagues = ['LaLiga Santander', 'Premier League', 'Bundesliga', 'Serie A TIM', 'Ligue 1 Conforama']
            #
            #     if league not in allowedLeagues:
            #         if lista == 1:
            #             lista = 2
            #         else:
            #             lista = 1
            #             cont += 1
            #         continue
            #
            #     query = League.objects.filter(name=league)
            #     if len(query) == 0:
            #         league = League.create(league)
            #     else:
            #         league = query[0]
            #     player.league = league
            #
            #     team = data[1]['data-original-title']
            #     team_img = data[1].find('img')['src']
            #     query = Team.objects.filter(name=team)
            #     if len(query) == 0:
            #         team = Team.create(team, league, team_img)
            #     else:
            #         team = query[0]
            #     player.team_origin = team
            #
            #     nation = data[2]['data-original-title']
            #     nation_img = data[2].find('img')['src']
            #     query = Nation.objects.filter(name=nation)
            #     if len(query) == 0:
            #         nation = Nation.create(nation, nation_img)
            #     else:
            #         nation = query[0]
            #     player.nation = nation
            #
            #     player.overall = ''.join(tds[1].findAll(text=True))
            #
            #     player.specific_position = ''.join(tds[2].findAll(text=True))
            #     # print(player.specific_position)
            #
            #     # dados['posicao'] = ''.join(tds[2].findAll(text=True))
            #     # dados['preco'] = ''.join(tds[4].findAll(text=True))
            #
            #     # rates = tds[7].findAll('span')
            #     # dados['attack_rate'] = ''.join(rates[0].findAll(text=True))
            #     # dados['defense_rate'] = ''.join(rates[1].findAll(text=True))
            #
            #     # dados['stats'] = dict()
            #     player.pace = ''.join(tds[8].findAll(text=True))
            #     player.shooting = ''.join(tds[9].findAll(text=True))
            #     player.passing = ''.join(tds[10].findAll(text=True))
            #     player.dribbling = ''.join(tds[11].findAll(text=True))
            #     player.defending = ''.join(tds[12].findAll(text=True))
            #     player.physical = ''.join(tds[13].findAll(text=True))
            #     player.likes = ''.join(tds[15].findAll(text=True))
            #
            #     # dados['altura'] = ''.join(tds[14].div.findAll(text=True))
            #
            #     index = get_jogador_index(player, position)
            #
            #     if index >= 0:
            #         other_player = Player.objects.get(id=index)
            #         if int(player.id) > other_player.id:
            #             other_player.delete()
            #             player.save()
            #         else:
            #             other_player.team_participant = None
            #             other_player.value = None
            #
            #             if int(player.likes) > other_player.likes:
            #                 other_player.likes = int(player.likes)
            #
            #             other_player.save()
            #     else:
            #         player.save()
            #
            #     if lista == 1:
            #         lista = 2
            #     else:
            #         lista = 1
            #         cont += 1

    driver.quit()

def getPlayersByPositionAlgorythm(position_id, n):
    n = int(n)
    players = list(Player.objects.filter(position=position_id).order_by('-overall', '-pace'))
    players = players[:n]
    # players.sort(key=lambda x: x.overall, reverse=True)
    # print(position_id, len(players), players)
    return players

def colourCell(cell, stat):
    redFill = PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
    orangeFill = PatternFill(start_color='FF8F00',end_color='FF8F00',fill_type='solid')
    yellowFill = PatternFill(start_color='F6EB0A',end_color='F6EB0A',fill_type='solid')
    lightGreenFill = PatternFill(start_color='2DD63A',end_color='2DD63A',fill_type='solid')
    greenFill = PatternFill(start_color='3E9C45',end_color='3E9C45',fill_type='solid')
    cell.value = stat

    if stat >= 90:
        cell.fill = greenFill
    elif stat >= 80:
        cell.fill = lightGreenFill
    elif stat >= 70:
        cell.fill = yellowFill
    elif stat >= 60:
        cell.fill = orangeFill
    else:
        cell.fill = redFill

    return cell

def getPlayersByPositionString(sheet, position, n):
    players = getPlayersByPositionAlgorythm(position.id, n)
    writePlayersSheet(sheet, players)
    return sheet

def writePlayersSheet (sheet, players):
    sheet.oddHeader.center.size = 14
    sheet.oddHeader.center.font = "Calibri,Bold"
    sheet.cell(row=1, column=1).value = 'PLAYER'
    sheet.cell(row=1, column=2).value = 'TEAM'
    sheet.cell(row=1, column=3).value = 'NATION'
    sheet.cell(row=1, column=4).value = 'POSITION'
    sheet.cell(row=1, column=5).value = 'OVERALL'
    sheet.cell(row=1, column=6).value = 'PACE'
    sheet.cell(row=1, column=7).value = 'SHOOTING'
    sheet.cell(row=1, column=8).value = 'PASSING'
    sheet.cell(row=1, column=9).value = 'DRIBBLING'
    sheet.cell(row=1, column=10).value = 'DEFENDING'
    sheet.cell(row=1, column=11).value = 'PHYSICAL'
    sheet.cell(row=1, column=12).value = 'VALUE'

    for i,player in enumerate(players):
        sheet.cell(row=i+2, column=1).value = player.name
        sheet.cell(row=i+2, column=2).value = player.team_origin.name
        sheet.cell(row=i+2, column=3).value = player.nation.name
        sheet.cell(row=i+2, column=4).value = player.specific_position
        sheet.cell(row=i+2, column=5).value = player.overall
        colourCell(sheet.cell(row=i+2, column=6), player.pace)
        colourCell(sheet.cell(row=i+2, column=7), player.shooting)
        colourCell(sheet.cell(row=i+2, column=8), player.passing)
        colourCell(sheet.cell(row=i+2, column=9), player.dribbling)
        colourCell(sheet.cell(row=i+2, column=10), player.defending)
        colourCell(sheet.cell(row=i+2, column=11), player.physical)

        # if player.value != None:
        sheet.cell(row=i+2, column=12).value = player.value

    return sheet